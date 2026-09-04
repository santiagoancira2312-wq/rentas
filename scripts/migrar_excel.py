#!/usr/bin/env python3
"""
Convierte el libro de Excel original en un archivo SQL de carga inicial.

    python3 scripts/migrar_excel.py <archivo.xlsx> > supabase/seed.sql

No inventa información: lo que el Excel no tiene (nombres de inquilinos,
teléfonos, contratos) se marca explícitamente para que el cliente lo complete.
Todo dato histórico se conserva, incluidos los pagos sin fecha real.
"""
import sys, re, unicodedata, datetime, collections
from pathlib import Path

import openpyxl

PROPIEDAD = "10000000-0000-0000-0000-000000000001"

# ── Identificadores estables: permiten volver a correr la migración sin duplicar
def uuid_de(prefijo: int, n: int) -> str:
    return f"{prefijo:08x}-0000-4000-8000-{n:012x}"

def slug(texto: str) -> str:
    t = unicodedata.normalize("NFKD", texto).encode("ascii", "ignore").decode()
    return re.sub(r"[^a-z0-9]+", "-", t.lower()).strip("-")

def sql(v) -> str:
    if v is None or v == "":
        return "null"
    if isinstance(v, bool):
        return "true" if v else "false"
    if isinstance(v, (int, float)):
        return repr(round(v, 4))
    if isinstance(v, datetime.datetime):
        return f"'{v.date().isoformat()}'"
    if isinstance(v, datetime.date):
        return f"'{v.isoformat()}'"
    return "'" + str(v).replace("'", "''") + "'"

def txt(v) -> str:
    """Texto que nunca es nulo: las columnas de notas son NOT NULL con default ''."""
    return "'" + str(v or "").replace("'", "''") + "'"

def fecha(v):
    return v.date() if isinstance(v, datetime.datetime) else v if isinstance(v, datetime.date) else None

# ─────────────────────────── Catálogo de unidades ───────────────────────────
# El Excel nombra los inmuebles como texto libre; aquí se les da tipo, tarifa y
# calendario de cobro reales. La tarifa semanal se deduce del propio archivo:
# los "esperados" mensuales de 5,500 / 4,400 son 5 y 4 lunes de $1,100.
UNIDADES = [
    # (excel,          nombre,             tipo,     frecuencia, tarifa, día, alta,      notas)
    ("Imprenta",       "Imprenta",         "local",  "monthly",  13250,  1, "2026-08-01", ""),
    ("SUSHI",          "Sushi",            "local",  "monthly",   8000,  1, "2026-08-01", ""),
    ("Dark Kitchen",   "Dark Kitchen",     "local",  "monthly",   5500,  1, "2026-08-01", ""),
    ("Bazar",          "Bazar",            "local",  "monthly",  10000, 15, "2026-08-01",
     "El archivo original no esperaba renta en diciembre: confirmar si el contrato termina en noviembre."),
    ("Cuarto 1",       "Cuarto 1",         "cuarto", "weekly",    1100,  1, "2026-08-01", ""),
    ("Cuarto 2",       "Cuarto 2",         "cuarto", "weekly",    1100,  1, "2026-08-01", ""),
    ("Cuarto 3",       "Cuarto 3",         "cuarto", "weekly",    1100,  1, "2026-08-01", ""),
    ("Cuarto 4",       "Cuarto 4",         "cuarto", "weekly",    1100,  1, "2026-08-01", ""),
    ("Cuarto 5",       "Cuarto 5",         "cuarto", "weekly",    1100,  1, "2026-08-01", ""),
    ("Cuarto 6",       "Cuarto 6",         "cuarto", "weekly",    1100,  1, "2026-08-01", ""),
    ("Cuarto 7 SB",    "Cuarto 7 SB",      "cuarto", "weekly",    1000,  1, "2026-08-01",
     "Ocupado por Billy, encargado de la propiedad. Su sueldo se registra aparte como egreso."),
    ("Cuarto 8 SB",    "Cuarto 8 SB",      "cuarto", "weekly",    1000,  1, "2026-08-01", ""),
    # Expansión planeada: se dan de alta hoy con fecha de arranque en diciembre,
    # así no afectan los indicadores de los meses anteriores.
    (None, "Airbnb 1",        "airbnb", "monthly", 5000, 1, "2026-12-01", "Expansión de diciembre de 2026."),
    (None, "Airbnb 2",        "airbnb", "monthly", 6500, 1, "2026-12-01", "Expansión de diciembre de 2026."),
    (None, "Airbnb 3",        "airbnb", "monthly", 6500, 1, "2026-12-01", "Expansión de diciembre de 2026."),
    (None, "Cuarto 3er piso", "cuarto", "monthly", 5000, 1, "2026-12-01", "Expansión de diciembre de 2026."),
]

TIPOS = [
    ("Cuarto",          "cuarto", "recurring", "bed",   1),
    ("Local comercial", "local",  "recurring", "store", 2),
    ("Airbnb",          "airbnb", "nightly",   "key",   3),
]

CATEGORIAS = [
    ("Agua", "agua", "#0ea5b7", 1), ("Luz", "luz", "#f59e0b", 2),
    ("Gas", "gas", "#8b5cf6", 3), ("Mantenimiento", "mantenimiento", "#64748b", 4),
    ("Reparaciones", "reparaciones", "#ef4444", 5), ("Limpieza", "limpieza", "#14b8a6", 6),
    ("Sueldos", "sueldos", "#3b82f6", 7), ("Impuestos", "impuestos", "#a855f7", 8),
    ("Servicios", "servicios", "#22c55e", 9), ("Otros", "otros", "#94a3b8", 10),
]

METODOS = [
    ("Efectivo", "efectivo", False, 1), ("Transferencia", "transferencia", True, 2),
    ("Depósito", "deposito", True, 3), ("Otro", "otro", False, 4),
]

# Egresos que en el archivo viven escritos a mano en el Resumen Ejecutivo,
# porque la tabla de captura quedó vacía.
EGRESOS = [
    ("2026-08-05", "luz",     "Recibo de luz de agosto",      1000),
    ("2026-08-05", "gas",     "Recarga de gas de agosto",      500),
    ("2026-08-10", "agua",    "Recibo de agua de agosto",     4527),
    ("2026-08-31", "sueldos", "Sueldo del encargado (Billy)", 4000),
]

TARIFAS_AGUA = [("2026-05-01", 81), ("2026-07-01", 101), ("2026-08-01", 93)]

RECIBOS_AGUA = [("2026-06-13", "2026-07-14", 54), ("2026-07-15", "2026-08-15", 47)]


def main() -> None:
    ruta = Path(sys.argv[1] if len(sys.argv) > 1 else "5_de_mayo.xlsx")
    wb = openpyxl.load_workbook(ruta, data_only=True)
    salida, avisos = [], []
    w = salida.append

    ids_tipo = {s: uuid_de(0x20, i) for i, (_, s, *_) in enumerate(TIPOS, 1)}
    ids_cat  = {s: uuid_de(0x60, i) for i, (_, s, *_) in enumerate(CATEGORIAS, 1)}
    ids_unidad, ids_inquilino, ids_lease = {}, {}, {}

    w("-- Generado por scripts/migrar_excel.py — no editar a mano.")
    w("-- Carga inicial a partir del libro de Excel original.")
    w("begin;\n")

    w("insert into properties (id, name, address, currency, timezone) values")
    w(f"  ({sql(PROPIEDAD)}, 'Casa 5 de Mayo', '', 'MXN', 'America/Monterrey')")
    w("on conflict (id) do nothing;\n")

    w("insert into unit_types (id, property_id, name, slug, billing_mode, icon, sort_order) values")
    w(",\n".join(
        f"  ({sql(ids_tipo[s])}, {sql(PROPIEDAD)}, {sql(n)}, {sql(s)}, {sql(m)}, {sql(ic)}, {o})"
        for n, s, m, ic, o in TIPOS) + "\non conflict (id) do nothing;\n")

    w("insert into expense_categories (id, property_id, name, slug, color, sort_order) values")
    w(",\n".join(
        f"  ({sql(ids_cat[s])}, {sql(PROPIEDAD)}, {sql(n)}, {sql(s)}, {sql(c)}, {o})"
        for n, s, c, o in CATEGORIAS) + "\non conflict (id) do nothing;\n")

    w("insert into payment_methods (id, property_id, name, slug, requires_reference, sort_order) values")
    w(",\n".join(
        f"  ({sql(uuid_de(0x70, i))}, {sql(PROPIEDAD)}, {sql(n)}, {sql(s)}, {sql(r)}, {o})"
        for i, (n, s, r, o) in enumerate(METODOS, 1)) + "\non conflict (id) do nothing;\n")

    # ── Unidades ──────────────────────────────────────────────────────────
    filas = []
    for i, (excel, nombre, tipo, frec, tarifa, dia, alta, notas) in enumerate(UNIDADES, 1):
        uid = uuid_de(0x30, i)
        ids_unidad[nombre] = uid
        if excel:
            ids_unidad[excel] = uid
        filas.append(f"  ({sql(uid)}, {sql(PROPIEDAD)}, {sql(ids_tipo[tipo])}, {sql(nombre)}, "
                     f"{sql(tarifa)}, {sql(frec)}, {dia}, {sql(alta)}, {txt(notas)})")
    w("insert into units (id, property_id, unit_type_id, name, base_rent, "
      "billing_frequency, billing_day, active_from, notes) values")
    w(",\n".join(filas) + "\non conflict (id) do nothing;\n")

    # ── Vencimientos y pagos del archivo ──────────────────────────────────
    ws = wb["Control de Rentas"]
    registros = [r for r in ws.iter_rows(min_row=11, max_row=206, values_only=True) if r[1]]

    ocupadas = {r[1] for r in registros if (r[4] or 0) > 0}
    desocupadas = {r[1] for r in registros} - ocupadas

    # El Excel no guarda inquilinos. Se crea uno por unidad con renta, marcado
    # para que el cliente lo complete, y así el historial queda enlazado.
    n = 0
    filas_inq, filas_lease = [], []
    for excel, nombre, tipo, frec, tarifa, dia, alta, _ in UNIDADES:
        if excel not in ocupadas:
            continue
        n += 1
        tid, lid = uuid_de(0x40, n), uuid_de(0x50, n)
        ids_inquilino[nombre] = tid
        ids_lease[nombre] = lid
        real = nombre == "Cuarto 7 SB"
        filas_inq.append(
            f"  ({sql(tid)}, {sql(PROPIEDAD)}, {sql('Billy' if real else f'Inquilino de {nombre}')}, "
            f"{txt('Encargado de la propiedad.' if real else 'PENDIENTE: capturar el nombre real del inquilino.')})")
        filas_lease.append(
            f"  ({sql(lid)}, {sql(ids_unidad[nombre])}, {sql(tid)}, '2026-08-01', "
            f"{sql(tarifa)}, {sql(frec)}, {dia}, 'Contrato reconstruido del archivo de Excel.')")

    w("-- El Excel no registra inquilinos: se crea uno por unidad ocupada para")
    w("-- enlazar el historial. El cliente captura los nombres reales al arrancar.")
    w("insert into tenants (id, property_id, full_name, notes) values")
    w(",\n".join(filas_inq) + "\non conflict (id) do nothing;\n")

    w("insert into leases (id, unit_id, tenant_id, starts_on, rent_amount, "
      "billing_frequency, billing_day, notes) values")
    w(",\n".join(filas_lease) + "\non conflict (id) do nothing;\n")

    w("update units set status = 'occupied' where id in (")
    w("  " + ", ".join(sql(ids_unidad[u[1]]) for u in UNIDADES if u[0] in ocupadas) + ");\n")

    # Cargos: se migran tal cual, con el importe que el archivo tenía en cada
    # vencimiento, para no alterar el histórico.
    filas_cargo, filas_pago = [], []
    sin_fecha = 0
    for i, r in enumerate(registros, 1):
        excel, frecuencia, limite = r[1], r[2], fecha(r[3])
        esperado, pago_en, pagado, situacion, obs = r[4] or 0, fecha(r[5]), r[6], r[9], r[10] or ""
        if not limite or excel not in ids_unidad:
            continue
        cid = uuid_de(0x80, i)
        semanal = (frecuencia or "").strip().lower().startswith("sem")
        ini, fin = (limite, limite + datetime.timedelta(days=6)) if semanal else (
            limite.replace(day=1),
            (limite.replace(day=28) + datetime.timedelta(days=4)).replace(day=1) - datetime.timedelta(days=1))
        nota = obs
        if situacion in ("Desocupado", "Billy") and not nota:
            nota = f"Situación en el archivo original: {situacion}."
        filas_cargo.append(
            f"  ({sql(cid)}, {sql(PROPIEDAD)}, {sql(ids_unidad[excel])}, "
            f"{sql(ids_lease.get(excel))}, {sql(ini)}, {sql(fin)}, {sql(limite)}, "
            f"{sql(esperado)}, 'Renta', {txt(nota)})")

        if pagado and pagado > 0:
            if pago_en is None:
                sin_fecha += 1
                pago_en = limite
                np = "Importe registrado en el archivo original sin fecha de pago; se asume la fecha límite."
            else:
                np = ""
            filas_pago.append(
                f"  ({sql(uuid_de(0x90, i))}, {sql(PROPIEDAD)}, {sql(cid)}, "
                f"{sql(ids_lease.get(excel))}, {sql(pago_en)}, {sql(pagado)}, {txt(np)})")

    w(f"-- {len(filas_cargo)} vencimientos de agosto a diciembre de 2026, tal como estaban en el archivo.")
    w("insert into charges (id, property_id, unit_id, lease_id, period_start, period_end, "
      "due_date, amount_expected, concept, notes) values")
    w(",\n".join(filas_cargo) + "\non conflict (id) do nothing;\n")

    w(f"-- {len(filas_pago)} pagos registrados ({sin_fecha} sin fecha real en el archivo).")
    w("insert into payments (id, property_id, charge_id, lease_id, paid_on, amount, notes) values")
    w(",\n".join(filas_pago) + "\non conflict (id) do nothing;\n")

    # ── Egresos ───────────────────────────────────────────────────────────
    w("-- Egresos reconstruidos del Resumen Ejecutivo: la tabla de captura del")
    w("-- archivo (tblEgresos) quedó vacía y su indicador apuntaba a otra columna.")
    w("insert into expenses (id, property_id, expense_category_id, incurred_on, concept, amount) values")
    w(",\n".join(
        f"  ({sql(uuid_de(0xa0, i))}, {sql(PROPIEDAD)}, {sql(ids_cat[c])}, {sql(f)}, {sql(cn)}, {sql(m)})"
        for i, (f, c, cn, m) in enumerate(EGRESOS, 1)) + "\non conflict (id) do nothing;\n")

    # ── Agua ──────────────────────────────────────────────────────────────
    w("insert into water_rates (id, property_id, effective_from, rate_per_m3) values")
    w(",\n".join(f"  ({sql(uuid_de(0xb0, i))}, {sql(PROPIEDAD)}, {sql(f)}, {sql(t)})"
                 for i, (f, t) in enumerate(TARIFAS_AGUA, 1))
      + "\non conflict (property_id, effective_from) do nothing;\n")

    wsa = wb["Control de Agua"]
    filas_agua, estimados = [], 0
    for i, r in enumerate(wsa.iter_rows(min_row=40, max_row=155, values_only=True), 1):
        f = fecha(r[2])
        if not f:
            continue
        am, pm, sig = r[3], r[4], r[5]
        if am is None and pm is None and sig is None:
            continue
        # En agosto el archivo rellenó días con fórmulas manuales; el consumo
        # queda idéntico (1.03) sin lecturas que lo respalden.
        estimado = am is None or sig is None
        estimados += estimado
        filas_agua.append(
            f"  ({sql(uuid_de(0xc0, i))}, {sql(PROPIEDAD)}, {sql(f)}, {sql(am)}, {sql(pm)}, "
            f"{sql(sig)}, {sql(estimado)}, "
            f"{txt('Lectura incompleta en el archivo original.' if estimado else '')})")

    w(f"-- {len(filas_agua)} días de bitácora ({estimados} incompletos, marcados como estimados).")
    w("insert into water_readings (id, property_id, read_on, reading_morning, reading_afternoon, "
      "reading_next_morning, is_estimated, notes) values")
    w(",\n".join(filas_agua) + "\non conflict (id) do nothing;\n")

    w("insert into water_bills (id, property_id, period_start, period_end, m3_billed) values")
    w(",\n".join(f"  ({sql(uuid_de(0xd0, i))}, {sql(PROPIEDAD)}, {sql(a)}, {sql(b)}, {sql(m)})"
                 for i, (a, b, m) in enumerate(RECIBOS_AGUA, 1)) + "\non conflict (id) do nothing;\n")

    w("commit;")

    # ── Reporte a stderr para que no contamine el SQL ────────────────────
    print("\n".join(salida))
    avisos = [
        f"unidades: {len(UNIDADES)} ({len(ocupadas)} con contrato reconstruido)",
        f"vencimientos: {len(filas_cargo)}",
        f"pagos: {len(filas_pago)} — {sin_fecha} sin fecha real, se asumió la fecha límite",
        f"egresos: {len(EGRESOS)} reconstruidos del Resumen Ejecutivo",
        f"lecturas de agua: {len(filas_agua)} — {estimados} incompletas",
        f"unidades sin renta en el archivo: {', '.join(sorted(desocupadas)) or 'ninguna'}",
        "PENDIENTE PARA EL CLIENTE: capturar nombres y teléfonos reales de los inquilinos.",
    ]
    print("\n".join("  · " + a for a in avisos), file=sys.stderr)


if __name__ == "__main__":
    main()
